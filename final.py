
#import libaries
import os
import openpyxl as xl
import math
from fpdf import FPDF


def generate():
    #load workbooks
    bookone = xl.load_workbook("static/pd.xlsx")
    booktwo = xl.load_workbook("static/et.xlsx")

    #load sheets
    pd = bookone['Sheet1']
    et = booktwo['Sheet1']


    #editing the current sheet, because of the spelling error 
    if et.cell(9,2).value == 'make my trip':
        et.cell(9,2).value = 'metro business'
        
    if et.cell(7,1).value == "Vinitha":
        et.cell(7,1).value = "Vinita"
        
    if et.cell(8,1).value == "Vinitha":
        et.cell(8,1).value = "Vinita"
        

    #takes max row
    pd_maxrow = pd.max_row
    et_maxrow = et.max_row


    #finding total working days set (8 hour work as one day)
    for row in range(2,et_maxrow+1):
        worked_cell = et.cell(row,3)
        if worked_cell.value:
            total_worked_days = math.ceil(int(worked_cell.value) / 8)
            total_worked_days_cell = et.cell(row,4)
            total_worked_days_cell.value = total_worked_days

            
    #remove empty rows
    for row in range(2,pd_maxrow):
        if not pd.cell(row,3).value:
            pd.delete_rows(row)
            
    for row in range(2,et_maxrow):
        if not et.cell(row,2).value:
            et.delete_rows(row)
            
    #set the max row again        
    pd_maxrow = pd.max_row
    et_maxrow = et.max_row


    #added values in column 1 of project details (fill the blanks)
    project_name = pd.cell(2,1)
    for row in range(2,pd_maxrow+1):
        if not pd.cell(row,1).value:
            pd.cell(row,1).value = project_name
        else:
            project_name = pd.cell(row,1).value
            
            
    ## function to remove space
    def remove(string): 
        return string.replace(" ", "")

    #remove space from project name and make it lower  (because the name is different in both xlsx file)
    for row in range(2,pd_maxrow):
        pd.cell(row,1).value = remove(pd.cell(row,1).value).lower()
        et.cell(row,2).value = remove(et.cell(row,2).value).lower()
        

    #function to convert the CAD to INR
    def conve(val):
        if val:
            if 'CAD' in val:
                val = 56.97 * int(val[4:])
                return("INR " + str(val))
            else:
                return val
        else:
            return ("INR 0")
        

    ## taking details
    project_name_list = []
    project_estimation_list = []
    other_expenses_list = []
    project_actual_cost = []

    for row in range(2,pd_maxrow+1):
        if pd.cell(row,2).value:
            project_name_list.append(pd.cell(row,1).value)
            project_actual_cost.append(0)
            
            exp = pd.cell(row,2).value
            project_estimation_list.append(conve(exp))
            
            oth = pd.cell(row,5).value
            other_expenses_list.append(conve(oth))
            
            

    #finding the cost and all values
    employee_cost = []
    employee_name = []
    for row in range(2,pd_maxrow):
        pd_project_name = pd.cell(row,1).value
        pd_employee_name = pd.cell(row,3).value
        
        for rows in range(2,et_maxrow+1):
            et_project_name = et.cell(rows,2).value
            et_employee_name = et.cell(rows,1).value
            
            if (pd_employee_name == et_employee_name and pd_project_name == et_project_name):
                days = et.cell(rows,4).value
                amount_per_day = pd.cell(row,4).value
                
                if pd_employee_name not in employee_name:
                    employee_name.append(pd_employee_name)
                    employee_cost.append(int(days)*int(amount_per_day))
                else:
                    index = employee_name.index(pd_employee_name)
                    employee_cost[index] += int(days)*int(amount_per_day)
                
                
                project_index = project_name_list.index(pd_project_name)
                project_actual_cost[project_index] += int(days)*int(amount_per_day)
                    


    #profit 
    profit_or_loss = []
    for i in range (len(project_actual_cost)):
        a = (float(project_estimation_list[i][4:]))
        project_estimation_list[i] = a
        b = (float(other_expenses_list[i][4:]))
        project_actual_cost[i] += b
        
        profit_or_loss.append(project_estimation_list[i] - project_actual_cost[i])

        
        
    #RESULTS
    print("RESULTS:")
    print("__________________________________________")
    print("Employee Revenue:")
    for i in range(len(employee_name)):
        print(employee_name[i] , " \t: " , employee_cost[i])

    print("__________________________________________")    
    print("Project :")
    print("project name \t actual revenue \t profit\loss")
    for i in range(len(project_actual_cost)):
        print(project_name_list[i],"\t",project_actual_cost[i],"\t\t",profit_or_loss[i])
    

    #####
    pdf = FPDF()
    pdf.add_page() 
    
    pdf.set_font("Arial", size = 25) 
    pdf.cell(200, 10, txt = "RESULTS", ln = 1, align = 'C') 
    pdf.set_font("Arial", size = 20) 
    pdf.cell(200, 10, txt = "Employee Revenue", ln = 2)

    pdf.set_font("Arial", size = 15) 
    for i in range(len(employee_name)):
        zz = str("         " + employee_name[i]) + "\t : " + str(employee_cost[i])
        pdf.cell(200, 10, txt = str(zz), ln = 3+i)

        
    pdf.set_font("Arial", size = 20)
    pdf.cell(200,10, txt = " ", ln = 3+i)
    pdf.cell(200,10,txt = "Project" , ln = 4+i)

    line = 5+i
    pdf.set_font("Arial", size = 15)
    for j in range(len(project_actual_cost)):
        zz = "         Project name  : " + str(project_name_list[j])
        pdf.cell(200,10, txt = str(zz), ln = line)
        line +=1
        
        zz = "         Actual Revenue: " + str(project_actual_cost[j])
        pdf.cell(200,10, txt = str(zz), ln = line)
        line +=1
        
        zz = "         Profit\loss   : " + str(profit_or_loss[j])
        pdf.cell(200,10, txt = str(zz), ln = line)
        line +=1
        if(profit_or_loss[j] > 0):
            pdf.cell(200,10, txt = "         Profit   :  YES", ln = line)
        else:
            pdf.cell(200,10, txt = "         Profit   :  NO", ln = line)
        
        line +=2
        pdf.cell(200,10, txt = " ", ln = line)

    pdf.output("Revenue.pdf")



    return output
